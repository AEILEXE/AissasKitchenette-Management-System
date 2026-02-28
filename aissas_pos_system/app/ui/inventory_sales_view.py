"""
app/ui/inventory_sales_view.py
────────────────────────────────
Part 4 — Sales page redesign.
- KPI summary cards: Total Sales, Total Orders, Average Order Value
- Chart wrapped in a styled card container
- Improved Daily / Monthly / Yearly toggle (pill-style radio buttons)
- Modern export buttons
- matplotlib logic is 100% unchanged
"""
from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import openpyxl
import os

from app.config import THEME
from app.db.database import Database
from app.db.dao import OrderDAO
from app.services.auth_service import AuthService
from app.ui import ui_scale
from app.utils import money


class InventorySalesView(tk.Frame):
    def __init__(self, parent: tk.Frame, db: Database, auth: AuthService):
        super().__init__(parent, bg=THEME["bg"])
        self.db         = db
        self.auth       = auth
        self.order_dao  = OrderDAO(db)

        self.var_view_type  = tk.StringVar(value="Daily")
        self.canvas_figure  = None

        self._build()
        self._refresh_data()

    # ──────────────────────────────────────────────────────────────────────────
    # Layout
    # ──────────────────────────────────────────────────────────────────────────

    def _build(self):
        f  = ui_scale.scale_font
        sp = ui_scale.s

        self.columnconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)   # chart row expands

        # ── Page header ───────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=THEME["bg"])
        hdr.grid(row=0, column=0, sticky="ew", padx=18, pady=(14, 10))
        hdr.columnconfigure(0, weight=1)

        tk.Label(
            hdr, text="Sales Analytics",
            bg=THEME["bg"], fg=THEME["text"],
            font=("Segoe UI", f(20), "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")

        tk.Label(
            hdr, text="Overview of completed order revenue",
            bg=THEME["bg"], fg=THEME["muted"],
            font=("Segoe UI", f(9)),
        ).grid(row=1, column=0, sticky="w")

        # View-type toggle (pill radio buttons)
        toggle_frame = tk.Frame(hdr, bg=THEME["beige"],
                                highlightthickness=1, highlightbackground=THEME["border"])
        toggle_frame.grid(row=0, column=1, rowspan=2, sticky="e")

        self._toggle_btns: dict[str, tk.Button] = {}
        for vt in ["Daily", "Monthly", "Yearly"]:
            btn = tk.Button(
                toggle_frame, text=vt,
                bg=THEME["beige"], fg=THEME["muted"],
                bd=0, padx=sp(12), pady=sp(6),
                cursor="hand2",
                font=("Segoe UI", f(9)),
                command=lambda v=vt: self._set_view_type(v),
            )
            btn.pack(side="left", padx=2, pady=2)
            self._toggle_btns[vt] = btn
        self._update_toggle_style()

        # ── KPI cards row ─────────────────────────────────────────────────────
        self.kpi_row = tk.Frame(self, bg=THEME["bg"])
        self.kpi_row.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))
        for i in range(3):
            self.kpi_row.columnconfigure(i, weight=1, uniform="kpi")

        # KPI placeholders (populated in _refresh_data)
        self._kpi_frames: list[tk.Frame] = []
        for i in range(3):
            pad = (0, 8) if i < 2 else (0, 0)
            card = tk.Frame(
                self.kpi_row, bg=THEME["panel"],
                highlightthickness=1, highlightbackground=THEME["border"],
            )
            card.grid(row=0, column=i, sticky="ew", padx=pad)
            self._kpi_frames.append(card)

        # ── Chart card ────────────────────────────────────────────────────────
        chart_card = tk.Frame(
            self, bg=THEME["panel"],
            highlightthickness=1, highlightbackground=THEME["border"],
        )
        chart_card.grid(row=3, column=0, sticky="nsew", padx=18, pady=(0, 10))
        chart_card.rowconfigure(1, weight=1)
        chart_card.columnconfigure(0, weight=1)

        # Chart header
        chart_hdr = tk.Frame(chart_card, bg=THEME["panel"])
        chart_hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(10, 6))
        chart_hdr.columnconfigure(0, weight=1)

        self._chart_title_lbl = tk.Label(
            chart_hdr, text="Sales Chart",
            bg=THEME["panel"], fg=THEME["text"],
            font=("Segoe UI", f(11), "bold"), anchor="w",
        )
        self._chart_title_lbl.grid(row=0, column=0, sticky="w")

        # Chart canvas container
        self.canvas_frame = tk.Frame(chart_card, bg="white")
        self.canvas_frame.grid(row=1, column=0, sticky="nsew", padx=2, pady=(0, 2))

        # ── Export bar ────────────────────────────────────────────────────────
        export_bar = tk.Frame(
            self, bg=THEME["panel"],
            highlightthickness=1, highlightbackground=THEME["border"],
        )
        export_bar.grid(row=4, column=0, sticky="ew", padx=18, pady=(0, 16))

        tk.Label(
            export_bar, text="Export",
            bg=THEME["panel"], fg=THEME["muted"],
            font=("Segoe UI", f(9), "bold"),
        ).pack(side="left", padx=(14, 10), pady=10)

        # separator
        tk.Frame(export_bar, bg=THEME["border"], width=1).pack(side="left", fill="y", pady=6)

        tk.Button(
            export_bar, text="⬇  Save as PDF",
            command=self._export_pdf,
            bg=THEME["brown"], fg="white",
            activebackground=THEME["brown_dark"], activeforeground="white",
            bd=0, padx=sp(14), pady=sp(8), cursor="hand2",
            font=("Segoe UI", f(9), "bold"),
        ).pack(side="left", padx=(12, 8), pady=8)

        tk.Button(
            export_bar, text="⬇  Save as Excel",
            command=self._export_excel,
            bg=THEME["accent"], fg="white",
            activebackground=THEME["brown"], activeforeground="white",
            bd=0, padx=sp(14), pady=sp(8), cursor="hand2",
            font=("Segoe UI", f(9), "bold"),
        ).pack(side="left", pady=8)

    # ──────────────────────────────────────────────────────────────────────────
    # Toggle helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _set_view_type(self, vt: str):
        self.var_view_type.set(vt)
        self._update_toggle_style()
        self._refresh_data()

    def _update_toggle_style(self):
        active = self.var_view_type.get()
        for vt, btn in self._toggle_btns.items():
            if vt == active:
                btn.configure(bg=THEME["brown"], fg="white",
                              font=("Segoe UI", ui_scale.scale_font(9), "bold"))
            else:
                btn.configure(bg=THEME["beige"], fg=THEME["muted"],
                              font=("Segoe UI", ui_scale.scale_font(9)))

    # ──────────────────────────────────────────────────────────────────────────
    # KPI cards
    # ──────────────────────────────────────────────────────────────────────────

    def _refresh_kpi(self, total_sales: float, order_count: int):
        f  = ui_scale.scale_font
        sp = ui_scale.s

        avg = total_sales / order_count if order_count else 0.0

        kpi_data = [
            ("Total Sales",       money(total_sales),   f"from {order_count} orders",   THEME["success"]),
            ("Total Orders",      str(order_count),     "completed orders",              THEME["brown"]),
            ("Average Order",     money(avg),           "per completed order",           THEME["accent"]),
        ]

        for frame, (title, value, subtitle, accent) in zip(self._kpi_frames, kpi_data):
            for w in frame.winfo_children():
                w.destroy()

            bar = tk.Frame(frame, bg=accent, height=4)
            bar.pack(fill="x")

            tk.Label(frame, text=title,
                     bg=THEME["panel"], fg=THEME["muted"],
                     font=("Segoe UI", f(9))).pack(anchor="w", padx=14, pady=(10, 2))
            tk.Label(frame, text=value,
                     bg=THEME["panel"], fg=accent,
                     font=("Segoe UI", f(18), "bold")).pack(anchor="w", padx=14)
            tk.Label(frame, text=subtitle,
                     bg=THEME["panel"], fg=THEME["muted"],
                     font=("Segoe UI", f(8))).pack(anchor="w", padx=14, pady=(2, 12))

    # ──────────────────────────────────────────────────────────────────────────
    # Data (unchanged logic)
    # ──────────────────────────────────────────────────────────────────────────

    def _get_sales_data(self):
        view_type = self.var_view_type.get()
        rows = self.order_dao.list_orders(status="Completed")
        if not rows:
            return [], 0.0, 0

        sales_dict: dict[str, float] = {}
        total_sales = 0.0
        order_count = 0

        for row in rows:
            try:
                dt_str = row["start_dt"]
                total  = float(row["total"])
                dt     = datetime.fromisoformat(dt_str) if isinstance(dt_str, str) else dt_str

                if view_type == "Daily":
                    key = dt.strftime("%Y-%m-%d")
                elif view_type == "Monthly":
                    key = dt.strftime("%Y-%m")
                else:
                    key = dt.strftime("%Y")

                sales_dict[key] = sales_dict.get(key, 0.0) + total
                total_sales    += total
                order_count    += 1
            except Exception:
                continue

        sorted_keys = sorted(sales_dict.keys())
        return [(k, sales_dict[k]) for k in sorted_keys], total_sales, order_count

    def _refresh_data(self):
        data, total_sales, order_count = self._get_sales_data()
        self._refresh_kpi(total_sales, order_count)

        # Update chart title
        vt = self.var_view_type.get()
        self._chart_title_lbl.configure(text=f"Sales — {vt} View")

        self._draw_graph(data)

    # ──────────────────────────────────────────────────────────────────────────
    # Chart (unchanged matplotlib logic)
    # ──────────────────────────────────────────────────────────────────────────

    def _draw_graph(self, data):
        for widget in self.canvas_frame.winfo_children():
            widget.destroy()

        if not data:
            tk.Label(
                self.canvas_frame,
                text="No sales data available",
                bg="white", fg=THEME["muted"],
                font=("Segoe UI", 12),
            ).pack(expand=True)
            return

        fig = Figure(figsize=(10, 5), dpi=80)
        ax  = fig.add_subplot(111)

        labels = [item[0] for item in data]
        values = [item[1] for item in data]

        bars = ax.bar(labels, values, color=THEME["brown"], edgecolor="none", linewidth=0)
        for bar in bars:
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2, height,
                f"₱{height:.0f}", ha="center", va="bottom", fontsize=9,
            )

        ax.set_xlabel("Period", fontsize=10)
        ax.set_ylabel("Sales (₱)", fontsize=10)
        ax.grid(axis="y", alpha=0.2)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        if len(labels) > 10:
            ax.tick_params(axis="x", rotation=45)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.canvas_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        self.canvas_figure = fig

    # ──────────────────────────────────────────────────────────────────────────
    # Export (unchanged logic)
    # ──────────────────────────────────────────────────────────────────────────

    def _export_pdf(self):
        data, total_sales, order_count = self._get_sales_data()
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"sales_{self.var_view_type.get()}.pdf",
        )
        if not file_path:
            return
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            from matplotlib.figure import Figure
            import matplotlib.gridspec as gridspec

            vt  = self.var_view_type.get()
            avg = total_sales / order_count if order_count else 0.0
            now = datetime.now().strftime("%Y-%m-%d %H:%M")

            with PdfPages(file_path) as pdf:
                fig = Figure(figsize=(10, 7))
                gs  = gridspec.GridSpec(2, 1, figure=fig, height_ratios=[1, 3],
                                        hspace=0.5)

                # ── Header / summary panel ────────────────────────────────────
                ax_info = fig.add_subplot(gs[0])
                ax_info.axis("off")

                summary_lines = [
                    f"Aissas Kitchenette — Sales Report ({vt})",
                    f"Generated: {now}",
                    f"",
                    f"Total Sales:     \u20b1{total_sales:,.2f}",
                    f"Total Orders:    {order_count}",
                    f"Average Order:   \u20b1{avg:,.2f}",
                ]
                if data:
                    summary_lines.append(f"Date Range:      {data[0][0]}  to  {data[-1][0]}")

                ax_info.text(
                    0.02, 0.95, "\n".join(summary_lines),
                    transform=ax_info.transAxes,
                    fontsize=10, verticalalignment="top",
                    fontfamily="monospace",
                    bbox=dict(boxstyle="round,pad=0.5", facecolor="#f9f0e8",
                              edgecolor="#c8a882", linewidth=1),
                )

                # ── Bar chart ─────────────────────────────────────────────────
                ax_chart = fig.add_subplot(gs[1])

                if data:
                    labels = [item[0] for item in data]
                    values = [item[1] for item in data]

                    bars = ax_chart.bar(labels, values, color="#6B4B3A",
                                        edgecolor="none")
                    for bar in bars:
                        h = bar.get_height()
                        if h > 0:
                            ax_chart.text(
                                bar.get_x() + bar.get_width() / 2, h,
                                f"\u20b1{h:,.0f}",
                                ha="center", va="bottom", fontsize=8,
                            )

                    ax_chart.set_xlabel("Period", fontsize=10)
                    ax_chart.set_ylabel("Sales (\u20b1)", fontsize=10)
                    ax_chart.set_title(f"Sales — {vt} View", fontsize=12, pad=10)
                    ax_chart.grid(axis="y", alpha=0.25)
                    ax_chart.spines["top"].set_visible(False)
                    ax_chart.spines["right"].set_visible(False)

                    if len(labels) > 10:
                        ax_chart.tick_params(axis="x", rotation=45)
                    fig.tight_layout(rect=[0, 0, 1, 1])
                else:
                    ax_chart.text(0.5, 0.5, "No sales data available",
                                  ha="center", va="center", fontsize=12,
                                  transform=ax_chart.transAxes)
                    ax_chart.axis("off")

                pdf.savefig(fig, bbox_inches="tight")

            messagebox.showinfo("PDF Exported", f"Saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("PDF Export Error", f"Failed to export PDF.\n\n{e}")

    def _export_excel(self):
        data, total_sales, order_count = self._get_sales_data()
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"sales_{self.var_view_type.get()}.xlsx",
        )
        if not file_path:
            return
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Report"

            vt  = self.var_view_type.get()
            avg = total_sales / order_count if order_count else 0.0
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            date_range = (
                f"{data[0][0]}  to  {data[-1][0]}" if data else "N/A"
            )

            # ── Title block ───────────────────────────────────────────────────
            title_fill   = PatternFill("solid", fgColor="6B4B3A")
            title_font   = Font(bold=True, size=14, color="FFFFFF")
            subtitle_font = Font(size=10, color="6E6E6E")
            bold_font    = Font(bold=True, size=10)
            header_fill  = PatternFill("solid", fgColor="EADFD2")
            header_font  = Font(bold=True, size=10, color="1F1F1F")
            total_fill   = PatternFill("solid", fgColor="F0FAF4")
            border_side  = Side(style="thin", color="D5C7B8")
            thin_border  = Border(bottom=border_side)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align  = Alignment(horizontal="right", vertical="center")

            ws.merge_cells("A1:C1")
            ws["A1"] = "Aissas Kitchenette — Sales Report"
            ws["A1"].font  = title_font
            ws["A1"].fill  = title_fill
            ws["A1"].alignment = center_align
            ws.row_dimensions[1].height = 28

            ws.merge_cells("A2:C2")
            ws["A2"] = f"View: {vt}  |  Generated: {now}  |  Date Range: {date_range}"
            ws["A2"].font      = subtitle_font
            ws["A2"].alignment = center_align
            ws.row_dimensions[2].height = 18

            # ── KPI summary row ───────────────────────────────────────────────
            ws.merge_cells("A3:C3")
            ws["A3"] = (
                f"Total Sales: \u20b1{total_sales:,.2f}     "
                f"Total Orders: {order_count}     "
                f"Average Order: \u20b1{avg:,.2f}"
            )
            ws["A3"].font      = bold_font
            ws["A3"].alignment = center_align
            ws["A3"].fill      = PatternFill("solid", fgColor="FFF3E0")
            ws.row_dimensions[3].height = 20

            # ── Column headers ────────────────────────────────────────────────
            headers = ["Period", "Sales (\u20b1)", "% of Total"]
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align if col_idx > 1 else Alignment(horizontal="left")
                cell.border    = thin_border
            ws.row_dimensions[5].height = 18

            # ── Data rows ─────────────────────────────────────────────────────
            for row_idx, (label, value) in enumerate(data, start=6):
                # Store as decimal fraction (e.g. 0.255) so Excel's 0.00% format
                # renders it correctly as "25.50%" instead of "2550.00%"
                pct = (value / total_sales) if total_sales else 0
                row_fill = PatternFill("solid", fgColor="FFFFFF" if (row_idx % 2 == 0) else "FAF7F4")

                c_period = ws.cell(row=row_idx, column=1, value=label)
                c_period.fill = row_fill

                c_sales = ws.cell(row=row_idx, column=2, value=value)
                c_sales.number_format = '"\u20b1"#,##0.00'
                c_sales.alignment     = right_align
                c_sales.fill          = row_fill

                c_pct = ws.cell(row=row_idx, column=3, value=round(pct, 6))
                c_pct.number_format = "0.00%"
                c_pct.alignment     = right_align
                c_pct.fill          = row_fill

            # ── Totals row ────────────────────────────────────────────────────
            totals_row = len(data) + 6
            ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True, size=10)
            ws.cell(row=totals_row, column=1).fill = total_fill

            c_total = ws.cell(row=totals_row, column=2, value=total_sales)
            c_total.number_format = '"\u20b1"#,##0.00'
            c_total.font          = Font(bold=True, size=10)
            c_total.alignment     = right_align
            c_total.fill          = total_fill

            c_total_pct = ws.cell(row=totals_row, column=3, value=1.0)
            c_total_pct.number_format = "0.00%"
            c_total_pct.font          = Font(bold=True, size=10)
            c_total_pct.alignment     = right_align
            c_total_pct.fill          = total_fill

            # ── Column widths ─────────────────────────────────────────────────
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 14

            # ── Freeze header rows ────────────────────────────────────────────
            ws.freeze_panes = "A6"

            wb.save(file_path)
            messagebox.showinfo("Excel Exported", f"Report saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Excel Export Error", f"Failed to export Excel.\n\n{e}")
